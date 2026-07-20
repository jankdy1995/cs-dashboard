#!/usr/bin/env python3
"""NeoTaste CS KPI Dashboard Builder.

Liest die aktuellste NeoTaste CS KPI-Excel und erzeugt ein self-contained
HTML-Dashboard (dashboard.html). Wird auch von der täglichen geplanten
Aufgabe verwendet.

Usage: python3 build_dashboard.py <xlsx-path> <output-html-path>
"""
import json
import sys
from datetime import date

import openpyxl

MONTHS_DE = ['', 'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
             'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']


def num(v):
    """Return float or None for messy cells (#REF!, '/', '#VALUE!', text)."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def extract(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    def rows(sheet, start=4, kwcol=0, maxcol=25):
        ws = wb[sheet]
        out = []
        for r in ws.iter_rows(min_row=start, max_row=ws.max_row,
                              max_col=maxcol, values_only=True):
            if num(r[kwcol]) is None:
                break
            out.append(r)
        return out

    d = {}

    hs = rows('📊 HubSpot_KPIs', maxcol=8)
    d['hubspot'] = [{
        'kw': int(r[0]), 'created': num(r[1]), 'user': num(r[2]),
        'partner': num(r[3]), 'messages': num(r[4]), 'csat': num(r[5]),
        'mtfr': num(r[6]), 'msg_per_ticket': num(r[7]),
    } for r in hs]

    ma = rows('🤖 MoinAI_KPIs', maxcol=11)
    d['moinai'] = [{
        'kw': int(r[0]), 'conv_hubspot': num(r[1]), 'conv_chatbot': num(r[2]),
        'conv_total': num(r[3]), 'share_chatbot': num(r[4]),
        'takeovers': num(r[5]), 'solved_bot': num(r[6]),
        'auto_bot': num(r[7]), 'auto_all': num(r[8]),
        'cost_per_ticket': num(r[9]), 'savings': num(r[10]),
    } for r in ma]

    tp = rows('👥 Team_Performance', maxcol=22)
    agents = {'Eli': (0, 1, 2, 3, 5), 'Jeanine': (6, 7, 8, 9, 11),
              'Vivien': (12, 13, 14, 15, 16), 'Jan': (17, 18, 19, 20, 21)}
    d['team'] = []
    for r in tp:
        row = {'kw': int(r[0])}
        for name, (kwc, msgc, ahtc, actc, notec) in agents.items():
            note = r[notec] if notec < len(r) else None
            note = str(note).strip() if note not in (None, '/', '') else None
            row[name] = {'messages': num(r[msgc]), 'aht': num(r[ahtc]),
                         'active_hours': num(r[actc]), 'note': note}
        d['team'].append(row)

    rf = rows('💸 Refund_Tracking', maxcol=9)
    d['refunds'] = [{
        'kw': int(r[0]), 'total': num(r[1]), 'refund_tickets': num(r[2]),
        'positive': num(r[3]), 'negative': num(r[4]),
        'decline_rate': num(r[5]), 'share': num(r[6]),
        'net': num(r[7]), 'gross': num(r[8]),
    } for r in rf]

    # CST_Costs: Spalten über Überschriften finden (Struktur ändert sich gelegentlich)
    ws = wb['💰CST_Costs']
    hdr2 = [c.value for c in ws[2]]  # z. B. '👤 CST Gesamt'
    hdr3 = [c.value for c in ws[3]]  # z. B. 'Created Tickets HubSpot', 'CPT'

    def find_col(headers, *needles):
        for i, h in enumerate(headers):
            if h and any(n.lower() in str(h).lower() for n in needles):
                return i
        return None

    col_total = find_col(hdr2, 'cst gesamt')
    col_fee = find_col(hdr2, 'moinai')
    col_hs = find_col(hdr3, 'created tickets hubspot')
    col_cb = find_col(hdr3, 'created tickets cb')
    # bevorzugt die Gesamtspalte "Created Tickets" (exakt), sonst HubSpot-Spalte
    col_tickets = next((i for i, h in enumerate(hdr3)
                        if str(h).strip().lower() == 'created tickets'), None)
    if col_tickets is None:
        col_tickets = find_col(hdr3, 'created tickets')
    col_cpt = find_col(hdr3, 'cpt')
    if col_total is None or col_tickets is None:
        raise ValueError('CST_Costs: Spalten "CST Gesamt"/"Created Tickets" '
                         'nicht gefunden — Sheet-Struktur prüfen!')

    d['costs'] = []
    for r in ws.iter_rows(min_row=4, max_row=ws.max_row,
                          max_col=ws.max_column, values_only=True):
        if r[0] is None:
            break
        total, tickets = num(r[col_total]), num(r[col_tickets])
        if total is None:  # Monat noch nicht befüllt
            continue
        cpt = num(r[col_cpt]) if col_cpt is not None else None
        if cpt is None and tickets:
            cpt = total / tickets
        d['costs'].append({
            'month': str(r[0]), 'total_cost': total, 'tickets': tickets,
            'cpt': cpt,
            'tickets_hs': num(r[col_hs]) if col_hs is not None else None,
            'tickets_cb': num(r[col_cb]) if col_cb is not None else None,
            'moinai_fee': num(r[col_fee]) if col_fee is not None else None,
        })

    d['keymetrics'] = build_keymetrics(d)
    return d


def build_keymetrics(d):
    """Monatsaggregate für die Key-Metrics-Sektion (letzter abgeschlossener
    Monat, Delta vs. Vormonat wo Daten vorhanden)."""
    year = date.today().year

    def month_of_kw(kw):
        # Woche gehört zum Monat ihres Donnerstags (ISO-Mehrheitsregel)
        try:
            return date.fromisocalendar(year, kw, 4).month
        except ValueError:
            return None

    def group(rows_):
        g = {}
        for r in rows_:
            m = month_of_kw(r['kw'])
            if m:
                g.setdefault(m, []).append(r)
        return g

    ghs, gma = group(d['hubspot']), group(d['moinai'])
    grf, gtm = group(d['refunds']), group(d['team'])
    months = sorted(ghs)
    if not months:
        return None
    cur_month_today = date.today().month
    completed = [m for m in months if m < cur_month_today]
    sel = completed[-1] if completed else months[-1]

    def avg(vals):
        vals = [v for v in vals if v is not None]
        return sum(vals) / len(vals) if vals else None

    def total(vals):
        vals = [v for v in vals if v is not None]
        return sum(vals) if vals else None

    costs_by_month = {c['month']: c for c in d['costs']}

    def km(m):
        if m < 1:
            return None
        name = MONTHS_DE[m]
        hsr, mar, rfr, tmr = (g.get(m, []) for g in (ghs, gma, grf, gtm))
        c = costs_by_month.get(name)
        if not hsr and not c:
            return None
        wsum = wmsg = 0
        for r in tmr:
            for a in ('Eli', 'Jeanine', 'Vivien', 'Jan'):
                mm, ah = r[a]['messages'], r[a]['aht']
                if mm and ah:
                    wsum += mm * ah
                    wmsg += mm
        th = c['tickets_hs'] if c else None
        tc = c['tickets_cb'] if c else None
        return {
            'month': name,
            'tickets_hs': th, 'tickets_cb': tc,
            'tickets_total': (th or 0) + (tc or 0) if (th or tc) else None,
            'mtfr': avg([r['mtfr'] for r in hsr]),
            'csat': avg([r['csat'] for r in hsr]),
            'auto_all': avg([r['auto_all'] for r in mar]),
            'savings_rdr': total([r['net'] for r in rfr]),
            'cpt': c['cpt'] if c else None,
            'aht': wsum / wmsg if wmsg else None,
        }

    return {'cur': km(sel), 'prev': km(sel - 1),
            'prev_month': MONTHS_DE[sel - 1] if sel > 1 else None}


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>NeoTaste · CS Dashboard</title>
<style>
  :root{
    --surface-1:#fcfcfb; --page:#f9f9f7;
    --ink-1:#0b0b0b; --ink-2:#52514e; --ink-3:#898781;
    --grid:#e1e0d9; --baseline:#c3c2b7;
    --border:rgba(11,11,11,.10);
    --s1:#2a78d6; --s2:#1baf7a; --s3:#eda100; --s4:#008300;
    --s5:#4a3aa7; --s6:#e34948; --s7:#e87ba4; --s8:#eb6834;
    --good:#006300; --bad:#d03b3b;
    --st-good:#0ca30c; --st-warn:#eda100; --st-crit:#d03b3b;
  }
  @media (prefers-color-scheme: dark){
    :root{
      --surface-1:#1a1a19; --page:#0d0d0d;
      --ink-1:#ffffff; --ink-2:#c3c2b7; --ink-3:#898781;
      --grid:#2c2c2a; --baseline:#383835;
      --border:rgba(255,255,255,.10);
      --s1:#3987e5; --s2:#199e70; --s3:#c98500; --s4:#008300;
      --s5:#9085e9; --s6:#e66767; --s7:#d55181; --s8:#d95926;
      --good:#0ca30c; --bad:#e66767;
    }
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--page);color:var(--ink-1);
    font:14px/1.45 system-ui,-apple-system,"Segoe UI",sans-serif;}
  .wrap{max-width:1240px;margin:0 auto;padding:28px 24px 48px;}
  header{display:flex;justify-content:space-between;align-items:flex-start;gap:16px;}
  header h1{font-size:21px;margin:0 0 2px;font-weight:650;}
  header .sub{color:var(--ink-3);font-size:13px;}
  .reload-btn{flex:none;display:inline-flex;align-items:center;gap:7px;
    background:var(--surface-1);border:1px solid var(--border);border-radius:8px;
    color:var(--ink-2);font:13px/1 system-ui,-apple-system,"Segoe UI",sans-serif;
    padding:9px 14px;cursor:pointer;}
  .reload-btn:hover{color:var(--ink-1);border-color:var(--baseline);}
  .reload-btn svg{width:14px;height:14px;stroke:currentColor;}
  .reload-btn.spin svg{animation:rot .8s linear infinite;}
  @keyframes rot{to{transform:rotate(360deg)}}
  .section-title{font-size:15px;font-weight:650;margin:34px 0 12px;
    display:flex;align-items:center;gap:8px;}
  .section-title::before{content:'';flex:none;width:5px;height:16px;
    border-radius:3px;background:var(--acc,var(--baseline));}
  .section-title .rule{flex:1;height:1px;background:var(--grid);}
  /* Sektions-Akzente */
  #sec-hubspot .card{border-top:3px solid var(--s1);}
  #sec-moinai .card{border-top:3px solid var(--s2);}
  #sec-team .card{border-top:3px solid var(--s5);}
  #sec-refunds .card{border-top:3px solid var(--s8);}
  #sec-costs .card{border-top:3px solid var(--s3);}
  .hdr-controls{display:flex;gap:8px;flex:none;align-items:center;}
  .period-sel{background:var(--surface-1);border:1px solid var(--border);
    border-radius:8px;color:var(--ink-2);font:13px system-ui,-apple-system,"Segoe UI",sans-serif;
    padding:9px 10px;cursor:pointer;}
  .period-sel:hover{color:var(--ink-1);border-color:var(--baseline);}
  .tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(168px,1fr));
    gap:12px;margin-top:18px;}
  .tile{background:var(--surface-1);border:1px solid var(--border);
    border-radius:10px;padding:14px 16px;}
  .tile .label{color:var(--ink-2);font-size:12.5px;}
  .tile .value{font-size:26px;font-weight:600;margin-top:2px;}
  .tile .delta{font-size:12px;margin-top:3px;color:var(--ink-3);}
  .tile .delta.up{color:var(--good)} .tile .delta.down{color:var(--bad)}
  .tile .sub{color:var(--ink-3);font-size:12px;margin-top:2px;}
  .tile .goal{color:var(--ink-3);font-size:11px;margin-top:4px;}
  .tile.st-good{border-left:3px solid var(--st-good);}
  .tile.st-warn{border-left:3px solid var(--st-warn);}
  .tile.st-crit{border-left:3px solid var(--st-crit);}
  .tile .stdot{display:inline-block;width:9px;height:9px;border-radius:50%;
    margin-right:7px;vertical-align:2px;}
  .insights{background:var(--surface-1);border:1px solid var(--border);
    border-radius:10px;padding:14px 18px;margin-top:18px;}
  .insights h2{font-size:13.5px;font-weight:650;margin:0;display:flex;
    align-items:center;gap:8px;cursor:pointer;user-select:none;}
  .insights h2 .chev{margin-left:auto;color:var(--ink-3);font-size:11px;
    transition:transform .15s;}
  .insights.open h2 .chev{transform:rotate(180deg);}
  .insights .body{display:none;margin-top:8px;}
  .insights.open .body{display:block;}
  .insights .count{background:var(--st-warn);color:#fff;border-radius:9px;
    font-size:11px;font-weight:650;padding:1px 8px;}
  .insights .row{display:flex;gap:9px;align-items:baseline;
    font-size:13px;margin:5px 0;color:var(--ink-2);}
  .insights .dot{flex:none;width:9px;height:9px;border-radius:50%;
    position:relative;top:1px;}
  .info-pop{position:absolute;z-index:8;background:var(--surface-1);
    border:1px solid var(--border);border-radius:8px;
    box-shadow:0 6px 18px rgba(0,0,0,.16);padding:10px 13px;
    font-size:12px;line-height:1.5;color:var(--ink-2);width:min(280px,90%);
    display:none;pointer-events:none;}
  .info-pop b{color:var(--ink-1);display:block;margin-bottom:3px;font-size:12px;}
  .explain-btn{background:none;border:none;color:var(--s1);font-size:12px;
    cursor:pointer;padding:0;margin:2px 0 6px;font-family:inherit;
    display:inline-flex;align-items:center;gap:5px;}
  .explain-btn .chev{font-size:9px;transition:transform .15s;}
  .explain-btn.open .chev{transform:rotate(180deg);}
  .explain-box{display:none;background:var(--page);border:1px solid var(--grid);
    border-radius:8px;padding:9px 12px;font-size:12px;line-height:1.5;
    color:var(--ink-2);margin:0 0 8px;}
  .explain-box.open{display:block;}
  .chev-km{color:var(--ink-3);font-size:11px;transition:transform .15s;}
  #kmTitleRow.open .chev-km{transform:rotate(180deg);}
  .tile .split{color:var(--ink-2);font-size:12px;margin-top:4px;}
  .km .tile .value{font-size:30px;}
  .km .tile{padding:16px 18px;}
  .grid{display:grid;grid-template-columns:repeat(12,1fr);gap:14px;}
  .card{background:var(--surface-1);border:1px solid var(--border);
    border-radius:10px;padding:16px 18px 12px;grid-column:span 6;
    min-width:0;}
  .card.w4{grid-column:span 4}.card.w12{grid-column:span 12}
  @media(max-width:900px){.card,.card.w4,.card.w12{grid-column:span 12}}
  .card h3{margin:0;font-size:13.5px;font-weight:650;}
  .card .hint{color:var(--ink-3);font-size:12px;margin:1px 0 8px;}
  .card-head{display:flex;justify-content:space-between;align-items:flex-start;}
  .view-sel{background:var(--surface-1);border:1px solid var(--border);
    border-radius:6px;color:var(--ink-2);font-size:11.5px;padding:3px 6px;
    cursor:pointer;font-family:inherit;}
  .view-sel:hover{color:var(--ink-1);border-color:var(--baseline);}
  .legend{display:flex;flex-wrap:wrap;gap:14px;margin:2px 0 6px;
    font-size:12px;color:var(--ink-2);}
  .legend .key{display:inline-flex;align-items:center;gap:6px;}
  .legend .swatch{width:12px;height:12px;border-radius:3px;}
  .legend .lkey{width:14px;height:2.5px;border-radius:2px;}
  .chart{position:relative;}
  .chart svg{display:block;width:100%;height:auto;}
  .tooltip{position:absolute;pointer-events:none;background:var(--surface-1);
    border:1px solid var(--border);border-radius:8px;
    box-shadow:0 4px 14px rgba(0,0,0,.14);padding:8px 11px;font-size:12px;
    min-width:130px;z-index:5;display:none;}
  .tooltip .tt-title{color:var(--ink-3);margin-bottom:4px;}
  .tooltip .tt-row{display:flex;align-items:center;gap:7px;margin:2px 0;}
  .tooltip .tt-key{width:12px;height:2.5px;border-radius:2px;flex:none;}
  .tooltip .tt-val{font-weight:650;margin-left:auto;padding-left:12px;}
  table.data{width:100%;border-collapse:collapse;font-size:12.5px;margin:6px 0 8px;}
  table.data th{color:var(--ink-3);font-weight:500;text-align:right;
    border-bottom:1px solid var(--grid);padding:4px 8px;}
  table.data td{text-align:right;padding:4px 8px;
    border-bottom:1px solid var(--grid);font-variant-numeric:tabular-nums;}
  table.data th:first-child,table.data td:first-child{text-align:left;}
  .foot{margin-top:30px;color:var(--ink-3);font-size:12px;}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <div>
      <h1>NeoTaste · Customer Success Dashboard</h1>
      <div class="sub" id="subline"></div>
    </div>
    <div class="hdr-controls">
      <select class="period-sel" id="periodSel" title="Zeitraum wählen">
        <option value="woche">Wöchentlich</option>
        <option value="monat">Monatlich</option>
        <option value="jahr">Jährlich</option>
      </select>
      <button class="reload-btn" id="reloadBtn" title="Aktuelle Zahlen laden">
        <svg viewBox="0 0 24 24" fill="none" stroke-width="2.2" stroke-linecap="round">
          <path d="M21 12a9 9 0 1 1-2.64-6.36M21 3v6h-6"/>
        </svg>
        Aktualisieren
      </button>
    </div>
  </header>

  <div class="insights" id="insights" style="display:none"></div>

  <div class="section-title" style="--acc:var(--s3);cursor:pointer;user-select:none" id="kmTitleRow">⭐ Key Metrics — <span id="km-title"></span><span class="rule"></span><span class="chev-km" id="kmChev">▼</span></div>
  <div class="tiles km" id="kmtiles" style="display:none"></div>

  <div class="section-title" style="--acc:var(--s1)">📆 Letzte Woche — <span id="week-title"></span><span class="rule"></span></div>
  <div class="tiles" id="tiles"></div>

  <div class="section-title" style="--acc:var(--s1)">🎫 Tickets &amp; Support (HubSpot)<span class="rule"></span></div>
  <div class="grid" id="sec-hubspot"></div>

  <div class="section-title" style="--acc:var(--s2)">🤖 Chatbot (MoinAI)<span class="rule"></span></div>
  <div class="grid" id="sec-moinai"></div>

  <div class="section-title" style="--acc:var(--s5)">👥 Team Performance<span class="rule"></span></div>
  <div class="grid" id="sec-team"></div>

  <div class="section-title" style="--acc:var(--s8)">💸 Refunds<span class="rule"></span></div>
  <div class="grid" id="sec-refunds"></div>

  <div class="section-title" style="--acc:var(--s3)">💰 Kosten<span class="rule"></span></div>
  <div class="grid" id="sec-costs"></div>

  <div class="foot" id="footline"></div>
</div>

<script>
const DATA = __DATA__;
const META = __META__;

const css = v => getComputedStyle(document.documentElement).getPropertyValue(v).trim();
const fmtN = v => v==null?'–':Math.round(v).toLocaleString('de-DE');
const fmtP = v => v==null?'–':(v*100).toLocaleString('de-DE',{maximumFractionDigits:1})+' %';
const fmtEuro = v => v==null?'–':Math.round(v).toLocaleString('de-DE')+' €';
const fmtEuro2 = v => v==null?'–':v.toLocaleString('de-DE',{minimumFractionDigits:2,maximumFractionDigits:2})+' €';
const fmtH = v => v==null?'–':v.toLocaleString('de-DE',{maximumFractionDigits:1})+' h';
const fmtMin = v => v==null?'–':v.toLocaleString('de-DE',{maximumFractionDigits:2})+' min';

/* ---------- Zielwerte & Ampel-Logik ---------- */
const TARGETS={
  csat:{dir:'up',good:0.71,warn:0.51,goal:'Ziel ≥ 71 %',line:0.71},
  mtfr:{dir:'down',good:24,warn:36,goal:'Ziel < 24 h',line:24},
  auto_all:{dir:'up',good:0.30,warn:-1,goal:'Ziel ≥ 30 %',line:0.30}, // unter Ziel = Gelb
  decline_rate:{dir:'up',good:0.90,warn:0.85,goal:'Ziel ≥ 90 %',line:0.90},
  cpt:{dir:'down',good:2.5,warn:3.5,goal:'Ziel ≤ 2,50 €',line:2.5},
  aht:{dir:'down',good:2.5,warn:3.5,goal:'Ziel ≤ 2,5 min',line:2.5},
};
function statusOf(key,v){
  const t=TARGETS[key];
  if(!t||v==null)return null;
  if(t.dir==='up')return v>=t.good?'good':(v>=t.warn?'warn':'crit');
  return v<=t.good?'good':(v<=t.warn?'warn':'crit');
}
const ST_COL={good:'--st-good',warn:'--st-warn',crit:'--st-crit'};

/* Info-Popup beim Hovern (Berechnungs-Erklärungen) */
function attachInfo(el,text,pos){
  el.style.position='relative';
  const pop=document.createElement('div');pop.className='info-pop';
  const b=document.createElement('b');
  b.textContent=text.startsWith('Hinweise:')?'ℹ️ Hinweise':'ℹ️ Wie wird das berechnet?';
  pop.append(b,document.createTextNode(text));
  if(pos==='below'){pop.style.left='0';pop.style.top='calc(100% + 6px)';}
  else{pop.style.right='14px';pop.style.top='46px';}
  el.append(pop);
  el.addEventListener('pointerenter',()=>{pop.style.display='block';});
  el.addEventListener('pointerleave',()=>{pop.style.display='none';});
}

/* ---------- Tiles ---------- */
function tile(label, value, delta, goodWhen, statusKey, statusVal, hoverInfo){
  const st=statusKey?statusOf(statusKey,statusVal):null;
  const el = document.createElement('div'); el.className='tile'+(st?' st-'+st:'');
  if(hoverInfo)attachInfo(el,hoverInfo,'below');
  const l = document.createElement('div'); l.className='label'; l.textContent=label;
  const v = document.createElement('div'); v.className='value';
  if(st){const dot=document.createElement('span');dot.className='stdot';
    dot.style.background=css(ST_COL[st]);v.append(dot);}
  v.append(document.createTextNode(value));
  el.append(l,v);
  if(delta!=null && isFinite(delta)){
    const d = document.createElement('div');
    const dir = delta>=0?'up':'down';
    let cls='';
    if(goodWhen) cls = (dir===goodWhen)?'up':'down';
    d.className='delta '+cls;
    const pct=(Math.abs(delta)*100).toLocaleString('de-DE',{maximumFractionDigits:1});
    d.textContent=(delta>=0?'▲ +':'▼ −')+pct+' % vs. Vorwoche';
    el.append(d);
  }
  if(statusKey&&TARGETS[statusKey]){const g=document.createElement('div');
    g.className='goal';g.textContent=TARGETS[statusKey].goal;el.append(g);}
  return el;
}
function rel(cur,prev){ return (cur!=null&&prev!=null&&prev!==0)?(cur-prev)/prev:null; }

function kmTile({label,value,delta,goodWhen,deltaLabel,sub,split,statusKey,statusVal}){
  const st=statusKey?statusOf(statusKey,statusVal):null;
  const el=document.createElement('div');el.className='tile'+(st?' st-'+st:'');
  const l=document.createElement('div');l.className='label';l.textContent=label;
  el.append(l);
  if(split){const sp=document.createElement('div');sp.className='split';
    sp.textContent=split;el.append(sp);}
  const v=document.createElement('div');v.className='value';
  if(st){const dot=document.createElement('span');dot.className='stdot';
    dot.style.background=css(ST_COL[st]);v.append(dot);}
  v.append(document.createTextNode(value));
  el.append(v);
  const d=document.createElement('div');
  if(delta!=null&&isFinite(delta)){
    const dir=delta>=0?'up':'down';
    d.className='delta '+(goodWhen?(dir===goodWhen?'up':'down'):'');
    const pct=(Math.abs(delta)*100).toLocaleString('de-DE',{maximumFractionDigits:1});
    d.textContent=(delta>=0?'↑ +':'↓ −')+pct+' % '+deltaLabel;
  }else{
    d.className='delta';d.textContent='Δ '+deltaLabel+': –';
  }
  el.append(d);
  if(sub){const s=document.createElement('div');s.className='sub';s.textContent=sub;el.append(s);}
  if(statusKey&&TARGETS[statusKey]){const g=document.createElement('div');
    g.className='goal';g.textContent=TARGETS[statusKey].goal;el.append(g);}
  return el;
}

/* ---------- Generic SVG charts ---------- */
const M={t:20,r:14,b:26,l:46}, W=560, H=240;

function makeSvg(){const s=document.createElementNS('http://www.w3.org/2000/svg','svg');
  s.setAttribute('viewBox',`0 0 ${W} ${H}`);return s;}
function line(s,x1,y1,x2,y2,stroke,w){const l=document.createElementNS(s.namespaceURI,'line');
  l.setAttribute('x1',x1);l.setAttribute('y1',y1);l.setAttribute('x2',x2);l.setAttribute('y2',y2);
  l.setAttribute('stroke',stroke);l.setAttribute('stroke-width',w||1);s.append(l);return l;}
function text(s,x,y,str,anchor,size,fill){const t=document.createElementNS(s.namespaceURI,'text');
  t.setAttribute('x',x);t.setAttribute('y',y);t.setAttribute('text-anchor',anchor||'start');
  t.setAttribute('font-size',size||10.5);t.setAttribute('fill',fill||css('--ink-3'));
  t.setAttribute('font-family','inherit');t.textContent=str;s.append(t);return t;}

function niceTicks(max){
  if(max<=0) max=1;
  const raw=max/4, mag=Math.pow(10,Math.floor(Math.log10(raw)));
  const step=[1,2,2.5,5,10].map(m=>m*mag).find(st=>max/st<=4.6)||10*mag;
  const ticks=[];for(let v=0;v<=max+1e-9;v+=step)ticks.push(v);
  if(ticks[ticks.length-1]<max)ticks.push(ticks[ticks.length-1]+step);
  return ticks;
}
function frame(svg,ymax,yFmt){
  const ticks=niceTicks(ymax), top=ticks[ticks.length-1];
  const y=v=>H-M.b-(v/top)*(H-M.t-M.b);
  ticks.forEach(v=>{
    line(svg,M.l,y(v),W-M.r,y(v),v===0?css('--baseline'):css('--grid'),1);
    text(svg,M.l-6,y(v)+3.5,yFmt(v),'end');
  });
  return y;
}

const VIEWS=[['linie','Linie'],['balken','Balken'],['flaeche','Fläche'],
             ['kreis','Kreis'],['tabelle','Tabelle']];
function chartCard(parent,{title,hint,width,legend,spec,table,hoverInfo,explain}){
  const card=document.createElement('div');card.className='card'+(width?' '+width:'');
  if(hoverInfo)attachInfo(card,hoverInfo);
  const head=document.createElement('div');head.className='card-head';
  const hwrap=document.createElement('div');
  const h=document.createElement('h3');h.textContent=title;hwrap.append(h);
  if(hint){const p=document.createElement('div');p.className='hint';p.textContent=hint;hwrap.append(p);}
  const sel=document.createElement('select');sel.className='view-sel';
  sel.title='Darstellung wählen';
  VIEWS.forEach(([v,lbl])=>{const o=document.createElement('option');
    o.value=v;o.textContent=lbl;sel.append(o);});
  head.append(hwrap,sel);card.append(head);
  if(explain){
    const btn=document.createElement('button');btn.className='explain-btn';
    const chev=document.createElement('span');chev.className='chev';chev.textContent='▼';
    btn.append(document.createTextNode('ℹ️ Wie wird das berechnet?'),chev);
    const boxE=document.createElement('div');boxE.className='explain-box';
    boxE.textContent=explain;
    btn.addEventListener('click',()=>{btn.classList.toggle('open');
      boxE.classList.toggle('open');});
    card.append(btn,boxE);
  }
  const legEl=document.createElement('div');legEl.className='legend';card.append(legEl);
  const chartEl=document.createElement('div');chartEl.className='chart';card.append(chartEl);
  const tblEl=document.createElement('div');tblEl.style.display='none';card.append(tblEl);
  buildTable(tblEl,table);

  // Ansicht aus URL wiederherstellen (teilbare Links)
  const slug='v-'+title.toLowerCase().replace(/[^a-z0-9äöüß]+/g,'-').replace(/^-+|-+$/g,'');
  const fromUrl=new URLSearchParams(location.search).get(slug);
  let view=VIEWS.some(([v])=>v===fromUrl)?fromUrl:(spec.defaultView||'linie');
  sel.value=view;

  function setLegend(items){
    legEl.replaceChildren();
    if(!items||items.length<2)return;
    items.forEach(k=>{const key=document.createElement('span');key.className='key';
      const sw=document.createElement('span');sw.className=k.type==='line'?'lkey':'swatch';
      sw.style.background=k.color;
      key.append(sw,document.createTextNode(k.name));legEl.append(key);});
  }
  function render(){
    chartEl.replaceChildren();
    chartEl.style.display='';tblEl.style.display='none';
    if(view==='tabelle'){chartEl.style.display='none';tblEl.style.display='';
      setLegend(null);return;}
    if(view==='kreis'){setLegend(pieChart(chartEl,spec));return;}
    setLegend(legend);
    if(view==='balken')barChart(chartEl,spec);
    else lineChart(chartEl,Object.assign({},spec,{area:view==='flaeche'}));
  }
  sel.addEventListener('change',()=>{
    view=sel.value;
    const u=new URL(location);u.searchParams.set(slug,view);
    history.replaceState(null,'',u);
    render();
  });
  render();
  parent.append(card);
}
function buildTable(el,{cols,rows}){
  const t=document.createElement('table');t.className='data';
  const tr=document.createElement('tr');
  cols.forEach(c=>{const th=document.createElement('th');th.textContent=c;tr.append(th);});
  t.append(tr);
  rows.forEach(r=>{const tr=document.createElement('tr');
    r.forEach(v=>{const td=document.createElement('td');td.textContent=v;tr.append(td);});
    t.append(tr);});
  el.append(t);
}
function tooltipFor(chartEl){
  const tt=document.createElement('div');tt.className='tooltip';chartEl.append(tt);
  return {
    show(px,py,title,rows){
      tt.replaceChildren();
      const h=document.createElement('div');h.className='tt-title';h.textContent=title;tt.append(h);
      rows.forEach(r=>{
        const row=document.createElement('div');row.className='tt-row';
        const k=document.createElement('span');k.className='tt-key';k.style.background=r.color;
        const n=document.createElement('span');n.textContent=r.name;
        const v=document.createElement('span');v.className='tt-val';v.textContent=r.value;
        row.append(k,n,v);tt.append(row);
      });
      tt.style.display='block';
      const rect=chartEl.getBoundingClientRect();
      let x=px+14,yv=py-10;
      if(x+tt.offsetWidth>rect.width) x=px-tt.offsetWidth-14;
      tt.style.left=Math.max(0,x)+'px';tt.style.top=Math.max(0,yv)+'px';
    },
    hide(){tt.style.display='none';}
  };
}

/* Line chart: series=[{name,color,values,fmt}], area: Flächen-Variante */
function lineChart(el,{labels,series,yFmt,labelSeries,area,target,trend}){
  const svg=makeSvg();el.append(svg);
  let max=Math.max(...series.flatMap(s=>s.values.filter(v=>v!=null)));
  if(target!=null)max=Math.max(max,target*1.08);
  const y=frame(svg,max,yFmt);
  if(target!=null){ // gestrichelte Ziellinie
    const tl=line(svg,M.l,y(target),W-M.r,y(target),css('--ink-3'),1);
    tl.setAttribute('stroke-dasharray','5 4');tl.setAttribute('opacity','.65');
    text(svg,W-M.r,y(target)-4,'Ziel','end',9.5);
  }
  if(trend&&series.length===1&&labels.length>=4){ // 4-Perioden-Trend
    const vals=series[0].values;
    const tv=vals.map((_,i)=>{
      const w=vals.slice(Math.max(0,i-3),i+1).filter(v=>v!=null);
      return w.length?w.reduce((a,b)=>a+b)/w.length:null;});
    const x0=i=>M.l+(labels.length===1?0.5:(i/(labels.length-1)))*(W-M.l-M.r);
    const d=tv.map((v,i)=>v==null?'':((i===0||tv[i-1]==null)?'M':'L')+x0(i)+' '+y(v)).join(' ');
    const tp=document.createElementNS(svg.namespaceURI,'path');
    tp.setAttribute('d',d);tp.setAttribute('fill','none');
    tp.setAttribute('stroke',css('--baseline'));tp.setAttribute('stroke-width',2);
    tp.setAttribute('opacity','.8');svg.append(tp);
  }
  const x=i=>M.l+(labels.length===1?0.5:(i/(labels.length-1)))*(W-M.l-M.r);
  labels.forEach((lb,i)=>text(svg,x(i),H-M.b+15,lb,'middle'));
  const cross=line(svg,0,M.t,0,H-M.b,css('--baseline'),1);cross.setAttribute('opacity','0');
  if(area)series.forEach(s=>{ // Flächen zuerst, damit Linien darüber liegen
    const pts=s.values.map((v,i)=>v==null?null:[x(i),y(v)]).filter(Boolean);
    if(pts.length<2)return;
    const d='M'+pts.map(p=>p[0]+' '+p[1]).join(' L ')+
      ` L${pts[pts.length-1][0]},${H-M.b} L${pts[0][0]},${H-M.b} Z`;
    const a=document.createElementNS(svg.namespaceURI,'path');
    a.setAttribute('d',d);a.setAttribute('fill',s.color);
    a.setAttribute('opacity','0.12');svg.append(a);
  });
  series.forEach(s=>{
    const pts=s.values.map((v,i)=>v==null?null:[x(i),y(v)]);
    const d=pts.map((p,i)=>p?((i===0||!pts[i-1])?'M':'L')+p[0]+' '+p[1]:'').join(' ');
    const path=document.createElementNS(svg.namespaceURI,'path');
    path.setAttribute('d',d);path.setAttribute('fill','none');
    path.setAttribute('stroke',s.color);path.setAttribute('stroke-width',2);
    path.setAttribute('stroke-linejoin','round');path.setAttribute('stroke-linecap','round');
    svg.append(path);
    pts.forEach(p=>{if(!p)return;
      const c=document.createElementNS(svg.namespaceURI,'circle');
      c.setAttribute('cx',p[0]);c.setAttribute('cy',p[1]);c.setAttribute('r',4);
      c.setAttribute('fill',s.color);c.setAttribute('stroke',css('--surface-1'));
      c.setAttribute('stroke-width',2);svg.append(c);});
  });
  // Werte an jeder KW: bei einer Serie diese, bei mehreren die Hauptserie
  const li=labelSeries!=null?labelSeries:0;
  const ls=series[Math.min(li,series.length-1)];
  ls.values.forEach((v,i)=>{
    if(v==null)return;
    const anchor=i===0?'start':(i===labels.length-1?'end':'middle');
    const dx=i===0?-4:(i===labels.length-1?4:0);
    text(svg,x(i)+dx,y(v)-9,(ls.fmt||yFmt)(v),anchor,10,css('--ink-1'))
      .setAttribute('font-weight','600');
  });
  const tt=tooltipFor(el);
  svg.addEventListener('pointermove',e=>{
    const r=svg.getBoundingClientRect(),sx=W/r.width;
    const mx=(e.clientX-r.left)*sx;
    let best=0,bd=1e9;
    labels.forEach((_,i)=>{const d=Math.abs(x(i)-mx);if(d<bd){bd=d;best=i;}});
    cross.setAttribute('x1',x(best));cross.setAttribute('x2',x(best));
    cross.setAttribute('opacity','.8');
    tt.show((e.clientX-r.left),(e.clientY-r.top),labels[best],
      series.map(s=>({name:s.name,color:s.color,
        value:(s.fmt||yFmt)(s.values[best])})));
  });
  svg.addEventListener('pointerleave',()=>{tt.hide();cross.setAttribute('opacity','0');});
}

/* Pie/Donut: mehrere Serien → Anteil je Serie (Summe über den Zeitraum);
   eine Serie → Anteil je Zeitpunkt (KW/Monat). Gibt Legenden-Einträge zurück. */
function pieChart(el,{labels,series,yFmt}){
  const PAL=['--s1','--s2','--s3','--s4','--s5','--s6','--s7','--s8'].map(css);
  const sum=v=>v.filter(x=>x!=null).reduce((a,b)=>a+b,0);
  let data;
  if(series.length>1){
    data=series.map(s=>({name:s.name,color:s.color,value:sum(s.values),
      fmt:s.fmt||yFmt}));
  }else{
    data=labels.map((lb,i)=>({name:lb,color:PAL[i%8],
      value:series[0].values[i],fmt:series[0].fmt||yFmt}));
    if(data.length>8){ // Farbslots nicht über 8 hinaus recyceln
      const rest=data.slice(7);
      data=data.slice(0,7);
      data.push({name:'Weitere',color:PAL[7],value:sum(rest.map(d=>d.value)),
        fmt:series[0].fmt||yFmt});
    }
  }
  data=data.filter(d=>d.value!=null&&d.value>0);
  const total=sum(data.map(d=>d.value));
  if(!total){el.textContent='Keine Daten';return [];}
  const svg=makeSvg();el.append(svg);
  const cx=W/2, cy=(H-6)/2, R=Math.min(W,H)/2-16, ri=R*0.62;
  const tt=tooltipFor(el);
  let a0=-Math.PI/2;
  data.forEach(d=>{
    const frac=d.value/total, a1=a0+frac*2*Math.PI;
    const p=(a,r)=>[cx+r*Math.cos(a),cy+r*Math.sin(a)];
    const [x0,y0]=p(a0,R),[x1,y1]=p(a1,R),[x2,y2]=p(a1,ri),[x3,y3]=p(a0,ri);
    const big=frac>0.5?1:0;
    const path=document.createElementNS(svg.namespaceURI,'path');
    path.setAttribute('d',`M${x0},${y0} A${R},${R} 0 ${big} 1 ${x1},${y1} `+
      `L${x2},${y2} A${ri},${ri} 0 ${big} 0 ${x3},${y3} Z`);
    path.setAttribute('fill',d.color);
    path.setAttribute('stroke',css('--surface-1'));   // 2px Surface-Gap
    path.setAttribute('stroke-width',2);
    path.addEventListener('pointermove',e=>{
      const rc=svg.getBoundingClientRect();
      tt.show(e.clientX-rc.left,e.clientY-rc.top,d.name,[
        {name:'Wert',color:d.color,value:d.fmt(d.value)},
        {name:'Anteil',color:'transparent',
         value:(frac*100).toLocaleString('de-DE',{maximumFractionDigits:1})+' %'}]);
    });
    path.addEventListener('pointerleave',()=>tt.hide());
    svg.append(path);
    if(frac>=0.06){ // nur größere Segmente direkt beschriften
      const [lx,ly]=p((a0+a1)/2,(R+ri)/2);
      text(svg,lx,ly+3.5,Math.round(frac*100)+' %','middle',10.5,'#ffffff')
        .setAttribute('font-weight','650');
    }
    a0=a1;
  });
  text(svg,cx,cy-2,yFmt(total),'middle',15,css('--ink-1'))
    .setAttribute('font-weight','650');
  text(svg,cx,cy+14,'Gesamt','middle',10.5);
  return data.map(d=>({name:d.name,color:d.color}));
}

/* Bar chart, optionally stacked: series=[{name,color,values,fmt}] */
function barChart(el,{labels,series,yFmt,totalLabelLast,target}){
  const svg=makeSvg();el.append(svg);
  const totals=labels.map((_,i)=>series.reduce((a,s)=>a+(s.values[i]||0),0));
  let bmax=Math.max(...totals);
  if(target!=null)bmax=Math.max(bmax,target*1.08);
  const y=frame(svg,bmax,yFmt);
  if(target!=null){
    const tl=line(svg,M.l,y(target),W-M.r,y(target),css('--ink-3'),1);
    tl.setAttribute('stroke-dasharray','5 4');tl.setAttribute('opacity','.65');
    text(svg,W-M.r,y(target)-4,'Ziel','end',9.5);
  }
  const band=(W-M.l-M.r)/labels.length, bw=Math.min(24,band*0.5);
  const surf=css('--surface-1');
  const tt=tooltipFor(el);
  labels.forEach((lb,i)=>{
    const cx=M.l+band*(i+0.5);
    text(svg,cx,H-M.b+15,lb,'middle');
    let y0=H-M.b;
    series.forEach((s,si)=>{
      const v=s.values[i]||0, h=(H-M.b)-y(v);
      const top=y0-h;
      const isTop=si===series.length-1;
      const r=document.createElementNS(svg.namespaceURI,'path');
      const x0=cx-bw/2, rad=isTop?4:0;
      const gap=si>0?2:0; // 2px Surface-Gap zwischen Segmenten
      const yTop=top+ (isTop?0:0), yBot=y0-gap;
      r.setAttribute('d',
        `M${x0},${yBot} L${x0},${yTop+rad}`+
        (rad?` Q${x0},${yTop} ${x0+rad},${yTop} L${x0+bw-rad},${yTop} Q${x0+bw},${yTop} ${x0+bw},${yTop+rad}`:
             ` L${x0+bw},${yTop}`)+
        ` L${x0+bw},${yBot} Z`);
      r.setAttribute('fill',s.color);
      svg.append(r);
      y0=top;
    });
    // Gesamtwert über jeder Säule
    text(svg,cx,y0-6,(series.length===1?(series[0].fmt||yFmt):fmtN)(totals[i]),
      'middle',10,css('--ink-1')).setAttribute('font-weight','600');
    // Hit target: ganze Band-Spalte
    const hit=document.createElementNS(svg.namespaceURI,'rect');
    hit.setAttribute('x',M.l+band*i);hit.setAttribute('y',M.t);
    hit.setAttribute('width',band);hit.setAttribute('height',H-M.t-M.b);
    hit.setAttribute('fill','transparent');
    hit.addEventListener('pointermove',e=>{
      const rc=svg.getBoundingClientRect();
      const rows=series.map(s=>({name:s.name,color:s.color,
        value:(s.fmt||yFmt)(s.values[i])}));
      if(series.length>1)rows.push({name:'Gesamt',color:'transparent',value:yFmt(totals[i])});
      tt.show(e.clientX-rc.left,e.clientY-rc.top,lb,rows);
    });
    hit.addEventListener('pointerleave',()=>tt.hide());
    svg.append(hit);
  });
}

/* ---------- Live-Daten aus Google Sheets ---------- */
const CSV_URLS={
  hubspot:'https://docs.google.com/spreadsheets/d/e/2PACX-1vSxbgk0YjSFwaYYGI6MPKUEmbu6sY4ew5tL4i6poTA3SbkU_MMb26j6TYaJ0ufZjozgl_ZpuYnxZBab/pub?gid=880512542&single=true&output=csv',
  moinai:'https://docs.google.com/spreadsheets/d/e/2PACX-1vSxbgk0YjSFwaYYGI6MPKUEmbu6sY4ew5tL4i6poTA3SbkU_MMb26j6TYaJ0ufZjozgl_ZpuYnxZBab/pub?gid=1229908884&single=true&output=csv',
  team:'https://docs.google.com/spreadsheets/d/e/2PACX-1vSxbgk0YjSFwaYYGI6MPKUEmbu6sY4ew5tL4i6poTA3SbkU_MMb26j6TYaJ0ufZjozgl_ZpuYnxZBab/pub?gid=25541729&single=true&output=csv',
  refunds:'https://docs.google.com/spreadsheets/d/e/2PACX-1vSxbgk0YjSFwaYYGI6MPKUEmbu6sY4ew5tL4i6poTA3SbkU_MMb26j6TYaJ0ufZjozgl_ZpuYnxZBab/pub?gid=1437991345&single=true&output=csv'
};

function parseCSV(text){
  const rows=[];let row=[],field='',inQ=false;
  for(let i=0;i<text.length;i++){
    const c=text[i];
    if(inQ){
      if(c==='"'){ if(text[i+1]==='"'){field+='"';i++;} else inQ=false; }
      else field+=c;
    }else if(c==='"')inQ=true;
    else if(c===','){row.push(field);field='';}
    else if(c==='\n'){row.push(field);rows.push(row);row=[];field='';}
    else if(c!=='\r')field+=c;
  }
  if(field!==''||row.length){row.push(field);rows.push(row);}
  return rows;
}
function parseNum(s){
  if(s==null)return null;
  s=String(s).trim();
  if(!s)return null;
  const pct=s.includes('%');
  s=s.replace(/[%€\s ]/g,'');
  if(/^-?\d{1,3}(\.\d{3})+(,\d+)?$/.test(s)) s=s.replace(/\./g,'').replace(',','.');
  else if(/^-?\d{1,3}(,\d{3})+(\.\d+)?$/.test(s)) s=s.replace(/,/g,'');
  else if(s.includes(',')&&!s.includes('.')) s=s.replace(',','.');
  const v=parseFloat(s);
  if(!isFinite(v))return null;
  return pct?v/100:v;
}
function dataRows(rows){ // Datenzeilen ab Zeile 4, Ende bei nicht-numerischer KW
  const out=[];
  for(let i=3;i<rows.length;i++){
    const kw=parseNum(rows[i][0]);
    if(kw==null||kw<1||kw>60)break;
    out.push(rows[i]);
  }
  return out;
}
function buildLiveData(hsT,maT,tmT,rfT){
  const n=parseNum;
  const hs=dataRows(parseCSV(hsT)).map(r=>({kw:Math.round(n(r[0])),created:n(r[1]),
    user:n(r[2]),partner:n(r[3]),messages:n(r[4]),csat:n(r[5]),mtfr:n(r[6]),
    msg_per_ticket:n(r[7])}));
  const ma=dataRows(parseCSV(maT)).map(r=>({kw:Math.round(n(r[0])),conv_hubspot:n(r[1]),
    conv_chatbot:n(r[2]),conv_total:n(r[3]),share_chatbot:n(r[4]),takeovers:n(r[5]),
    solved_bot:n(r[6]),auto_bot:n(r[7]),auto_all:n(r[8]),savings:n(r[10])}));
  const AG={Eli:[0,1,2,3,5],Jeanine:[6,7,8,9,11],Vivien:[12,13,14,15,16],Jan:[17,18,19,20,21]};
  const tm=dataRows(parseCSV(tmT)).map(r=>{
    const o={kw:Math.round(n(r[0]))};
    for(const[a,[,mc,ac,hc,nc]]of Object.entries(AG)){
      let note=(r[nc]||'').trim();
      if(note==='/'||note==='')note=null;
      o[a]={messages:n(r[mc]),aht:n(r[ac]),active_hours:n(r[hc]),note:note};
    }
    return o;});
  const rf=dataRows(parseCSV(rfT)).map(r=>({kw:Math.round(n(r[0])),total:n(r[1]),
    refund_tickets:n(r[2]),positive:n(r[3]),negative:n(r[4]),decline_rate:n(r[5]),
    share:n(r[6]),net:n(r[7]),gross:n(r[8])}));
  if(!hs.length)throw new Error('keine HubSpot-Daten');
  return {hubspot:hs,moinai:ma,team:tm,refunds:rf,costs:DATA.costs,live:true};
}
async function fetchLive(){
  const get=u=>fetch(u+'&_='+Date.now(),{cache:'no-store'})
    .then(r=>{if(!r.ok)throw new Error('HTTP '+r.status);return r.text();});
  const[a,b,c,d]=await Promise.all([get(CSV_URLS.hubspot),get(CSV_URLS.moinai),
    get(CSV_URLS.team),get(CSV_URLS.refunds)]);
  return buildLiveData(a,b,c,d);
}

/* ---------- Key Metrics client-seitig berechnen ---------- */
const MONTHS_DE=['','Januar','Februar','März','April','Mai','Juni','Juli',
  'August','September','Oktober','November','Dezember'];
function monthOfKW(kw){ // Monat des Donnerstags der ISO-Woche
  const y=new Date().getFullYear();
  const jan4=new Date(Date.UTC(y,0,4));
  const mon1=new Date(jan4);mon1.setUTCDate(jan4.getUTCDate()-((jan4.getUTCDay()+6)%7));
  const thu=new Date(mon1);thu.setUTCDate(mon1.getUTCDate()+(kw-1)*7+3);
  return thu.getUTCMonth()+1;
}
function computeKeymetrics(D){
  const grp=rows=>{const g={};rows.forEach(r=>{const m=monthOfKW(r.kw);
    (g[m]=g[m]||[]).push(r);});return g;};
  const ghs=grp(D.hubspot),gma=grp(D.moinai),grf=grp(D.refunds),gtm=grp(D.team);
  const months=Object.keys(ghs).map(Number).sort((a,b)=>a-b);
  if(!months.length)return null;
  const nowM=new Date().getMonth()+1;
  const done=months.filter(m=>m<nowM);
  const sel=done.length?done[done.length-1]:months[months.length-1];
  const avg=v=>{v=v.filter(x=>x!=null);return v.length?v.reduce((a,b)=>a+b)/v.length:null;};
  const sum=v=>{v=v.filter(x=>x!=null);return v.length?v.reduce((a,b)=>a+b):null;};
  const costsBy={};(D.costs||[]).forEach(c=>costsBy[c.month]=c);
  function km(m){
    if(m<1)return null;
    const name=MONTHS_DE[m],hsr=ghs[m]||[],mar=gma[m]||[],rfr=grf[m]||[],tmr=gtm[m]||[];
    const c=costsBy[name];
    if(!hsr.length&&!c)return null;
    let ws=0,wm=0;
    tmr.forEach(r=>['Eli','Jeanine','Vivien','Jan'].forEach(a=>{
      const mm=r[a].messages,ah=r[a].aht;
      if(mm&&ah){ws+=mm*ah;wm+=mm;}}));
    const th=c?c.tickets_hs:null,tc=c?c.tickets_cb:null;
    return {month:name,tickets_hs:th,tickets_cb:tc,
      tickets_total:(th||tc)?(th||0)+(tc||0):null,
      mtfr:avg(hsr.map(r=>r.mtfr)),csat:avg(hsr.map(r=>r.csat)),
      auto_all:avg(mar.map(r=>r.auto_all)),savings_rdr:sum(rfr.map(r=>r.net)),
      cpt:c?c.cpt:null,aht:wm?ws/wm:null};
  }
  return {cur:km(sel),prev:km(sel-1),prev_month:sel>1?MONTHS_DE[sel-1]:null};
}

/* ---------- Zeitraum-Aggregation (Woche/Monat/Jahr) ---------- */
function aggregate(D,period){
  if(period==='woche'){
    const lab=rows=>rows.map(r=>Object.assign({},r,{label:'KW '+r.kw}));
    return Object.assign({},D,{hubspot:lab(D.hubspot),moinai:lab(D.moinai),
      team:lab(D.team),refunds:lab(D.refunds)});
  }
  const y=new Date().getFullYear();
  const key=r=>period==='jahr'?String(y):MONTHS_DE[monthOfKW(r.kw)];
  function agg(rows,how){
    const order=[],map={};
    rows.forEach(r=>{const k=key(r);
      if(!map[k]){map[k]={label:k,_rows:[]};order.push(map[k]);}
      map[k]._rows.push(r);});
    return order.map(b=>{
      const o={label:b.label};
      for(const[f,h]of Object.entries(how)){
        const v=b._rows.map(r=>r[f]).filter(x=>x!=null);
        o[f]=v.length?(h==='sum'?v.reduce((a,c)=>a+c):v.reduce((a,c)=>a+c)/v.length):null;
      }
      return o;});
  }
  const hs=agg(D.hubspot,{created:'sum',user:'sum',partner:'sum',
    messages:'sum',csat:'avg',mtfr:'avg'});
  const ma=agg(D.moinai,{conv_hubspot:'sum',conv_chatbot:'sum',conv_total:'sum',
    takeovers:'sum',auto_bot:'avg',auto_all:'avg',savings:'sum'});
  const rf=agg(D.refunds,{refund_tickets:'sum',positive:'sum',negative:'sum',
    decline_rate:'avg',share:'avg',net:'sum',gross:'sum'});
  const order=[],map={};
  D.team.forEach(r=>{const k=key(r);
    if(!map[k]){map[k]={label:k,rows:[]};order.push(map[k]);}
    map[k].rows.push(r);});
  const tm=order.map(b=>{
    const o={label:b.label};
    ['Eli','Jeanine','Vivien','Jan'].forEach(a=>{
      const ms=b.rows.map(r=>r[a].messages).filter(v=>v!=null);
      const ah=b.rows.map(r=>r[a].aht).filter(v=>v!=null);
      const hh=b.rows.map(r=>r[a].active_hours).filter(v=>v!=null);
      const nt=b.rows.map(r=>r[a].note).filter(Boolean);
      o[a]={messages:ms.length?ms.reduce((x,z)=>x+z):null,
            aht:ah.length?ah.reduce((x,z)=>x+z)/ah.length:null,
            active_hours:hh.length?hh.reduce((x,z)=>x+z):null,
            note:nt.length?nt.join(', '):null};});
    return o;});
  return Object.assign({},D,{hubspot:hs,moinai:ma,team:tm,refunds:rf});
}

/* ---------- Build ---------- */
let lastData=null;
function renderAll(D){
['kmtiles','tiles','sec-hubspot','sec-moinai','sec-team','sec-refunds','sec-costs']
  .forEach(id=>document.getElementById(id).replaceChildren());
lastData=D;
const period=document.getElementById('periodSel').value;
const A=aggregate(D,period);
const hs=A.hubspot, ma=A.moinai, tm=A.team, rf=A.refunds, co=D.costs;
const kwl=hs.map(r=>r.label);
// Kacheln "Aktuelle Woche" + Deltas immer auf Wochenbasis (Rohdaten)
const curW=D.hubspot[D.hubspot.length-1];
const cur=Object.assign({},curW,{label:'KW '+curW.kw});
const prev=D.hubspot[D.hubspot.length-2]||{};
const maC=D.moinai[D.moinai.length-1], maP=D.moinai[D.moinai.length-2]||{};
const rfC=D.refunds[D.refunds.length-1], rfP=D.refunds[D.refunds.length-2]||{};
const coC=co[co.length-1]||{};

const standTxt=D.live
  ?'Live aus Google Sheets · '+new Date().toLocaleString('de-DE',{dateStyle:'short',timeStyle:'short'})+' Uhr'
  :`Stand: ${META.updated} (letzter Build)`;
document.getElementById('subline').textContent=
  `Letzte Woche: KW ${cur.kw} · ${standTxt}`;
document.getElementById('footline').textContent=
  `Quelle: NeoTaste CS KPI-Liste (Google Sheets) · ${standTxt} · Zeitraum: KW ${D.hubspot[0].kw}–${cur.kw}`;

/* Aufmerksamkeit: automatische Abweichungs-Hinweise */
(function(){
  const box=document.getElementById('insights');
  const items=[];
  const RW=D.hubspot,RM=D.moinai,RR=D.refunds;
  const lastOf=a=>a[a.length-1],prevOf=a=>a[a.length-2];
  function streak(rows,f,key){
    let n=0;
    for(let i=rows.length-1;i>=0;i--){
      const st=statusOf(key,rows[i][f]);
      if(st==='warn'||st==='crit')n++;else break;
    }
    return n;
  }
  const streaks=[
    [RW,'csat','csat','CSAT',v=>fmtP(v)],
    [RW,'mtfr','mtfr','Median First Reply Time',v=>fmtH(v)],
    [RM,'auto_all','auto_all','Automation Rate',v=>fmtP(v)],
    [RR,'decline_rate','decline_rate','Refund Decline Rate',v=>fmtP(v)],
  ];
  streaks.forEach(([rows,f,key,name,fmt])=>{
    const n=streak(rows,f,key);
    const cv=lastOf(rows)[f];
    if(n>=2)items.push({sev:3,st:statusOf(key,cv),
      txt:name+' verfehlt das Ziel seit '+n+' Wochen (aktuell '+fmt(cv)+' · '+TARGETS[key].goal.replace('Ziel ','Ziel: ')+').'});
    else if(n===1)items.push({sev:1,st:statusOf(key,cv),
      txt:name+' liegt diese Woche unter Ziel ('+fmt(cv)+' · '+TARGETS[key].goal.replace('Ziel ','Ziel: ')+').'});
  });
  const moves=[
    [RW,'created','Erstellte Tickets','down',fmtN],
    [RW,'mtfr','Median First Reply Time','down',fmtH],
    [RM,'auto_all','Automation Rate','up',fmtP],
    [RM,'savings','Weekly Savings','up',fmtEuro],
    [RR,'net','Net Contribution','up',fmtEuro],
  ];
  moves.forEach(([rows,f,name,goodDir,fmt])=>{
    const c=lastOf(rows),pv=prevOf(rows);
    if(!c||!pv||c[f]==null||!pv[f])return;
    const d=(c[f]-pv[f])/pv[f];
    if(Math.abs(d)<0.10)return;
    const good=(d>=0)===(goodDir==='up');
    items.push({sev:2,st:good?'good':'warn',
      txt:name+' '+(d>=0?'+':'−')+(Math.abs(d)*100).toLocaleString('de-DE',{maximumFractionDigits:0})+' % vs. Vorwoche ('+fmt(c[f])+').'});
  });
  const created=RW.map(r=>r.created).filter(v=>v!=null);
  if(created.length>=3&&lastOf(RW).created===Math.max(...created))
    items.push({sev:1,st:'warn',txt:'Erstellte Tickets erreichen mit '+fmtN(lastOf(RW).created)+' den höchsten Wochenwert seit Beginn der Aufzeichnung.'});
  items.sort((a,b)=>b.sev-a.sev);
  const top=items.slice(0,5);
  box.replaceChildren();
  if(!top.length){box.style.display='none';}
  else{
    box.style.display='';
    box.classList.toggle('open',window.__insightsOpen===true);
    const h=document.createElement('h2');
    const title=document.createElement('span');
    title.textContent='🔎 Aufmerksamkeit — KW '+lastOf(RW).kw;
    const cnt=document.createElement('span');cnt.className='count';
    cnt.textContent=top.length+(top.length===1?' Hinweis':' Hinweise');
    const chev=document.createElement('span');chev.className='chev';chev.textContent='▼';
    h.append(title,cnt,chev);
    h.addEventListener('click',()=>{
      window.__insightsOpen=!box.classList.contains('open');
      box.classList.toggle('open');
    });
    const body=document.createElement('div');body.className='body';
    top.forEach(it=>{
      const row=document.createElement('div');row.className='row';
      const dot=document.createElement('span');dot.className='dot';
      dot.style.background=css(ST_COL[it.st]||'--st-warn');
      const t=document.createElement('span');t.textContent=it.txt;
      row.append(dot,t);body.append(row);
    });
    box.append(h,body);
  }
})();

/* Key Metrics (Monatssicht) */
const KM=computeKeymetrics(D);
if(KM&&KM.cur){
  const k=KM.cur,p=KM.prev||{};
  const vsLbl='vs. '+(KM.prev_month||'Vormonat');
  document.getElementById('km-title').textContent=
    k.month+' · Δ '+vsLbl;
  const kmEl=document.getElementById('kmtiles');
  const t=(o)=>kmEl.append(kmTile(o));
  t({label:'Ticket Volume',value:fmtN(k.tickets_total),
     split:(k.tickets_hs!=null?fmtN(k.tickets_hs)+' HS':'')+
           (k.tickets_cb!=null?' · '+fmtN(k.tickets_cb)+' CB':''),
     delta:rel(k.tickets_total,p.tickets_total),goodWhen:'down',
     deltaLabel:vsLbl,sub:'Tickets gesamt (HubSpot + Chatbot)'});
  t({label:'Median First Reply Time',value:fmtH(k.mtfr),
     delta:rel(k.mtfr,p.mtfr),goodWhen:'down',deltaLabel:vsLbl,
     sub:'Ø der Wochen-Mediane',statusKey:'mtfr',statusVal:k.mtfr});
  t({label:'CSAT',value:fmtP(k.csat),
     delta:rel(k.csat,p.csat),goodWhen:'up',deltaLabel:vsLbl,
     sub:'zufriedene Bewertungen',statusKey:'csat',statusVal:k.csat});
  t({label:'Automation Rate',value:fmtP(k.auto_all),
     delta:rel(k.auto_all,p.auto_all),goodWhen:'up',deltaLabel:vsLbl,
     sub:'aller Tickets',statusKey:'auto_all',statusVal:k.auto_all});
  t({label:'Savings by RDR',value:fmtEuro(k.savings_rdr),
     delta:rel(k.savings_rdr,p.savings_rdr),goodWhen:'up',deltaLabel:vsLbl,
     sub:'Net Money Contribution'});
  t({label:'Cost per Ticket',value:fmtEuro2(k.cpt),
     delta:rel(k.cpt,p.cpt),goodWhen:'down',deltaLabel:vsLbl,
     sub:'inkl. Chatbot-Tickets',statusKey:'cpt',statusVal:k.cpt});
  t({label:'Average Handling Time',value:fmtMin(k.aht),
     delta:rel(k.aht,p.aht),goodWhen:'down',deltaLabel:vsLbl,
     sub:'Ø gewichtet über alle Agents',statusKey:'aht',statusVal:k.aht});
}
document.getElementById('kmtiles').style.display=window.__kmOpen?'':'none';
document.getElementById('kmTitleRow').classList.toggle('open',!!window.__kmOpen);

document.getElementById('week-title').textContent='KW '+cur.kw;
const tiles=document.getElementById('tiles');
tiles.append(
  tile('Erstellte Tickets (KW '+cur.kw+')',fmtN(cur.created),rel(cur.created,prev.created),'down'),
  tile('CSAT',fmtP(cur.csat),rel(cur.csat,prev.csat),'up','csat',cur.csat),
  tile('Median First Reply Time',fmtH(cur.mtfr),rel(cur.mtfr,prev.mtfr),'down','mtfr',cur.mtfr),
  tile('Automation Rate (alle Tickets)',fmtP(maC.auto_all),rel(maC.auto_all,maP.auto_all),'up','auto_all',maC.auto_all),
  tile('Refund Decline Rate',fmtP(rfC.decline_rate),rel(rfC.decline_rate,rfP.decline_rate),'up','decline_rate',rfC.decline_rate),
  tile('Weekly Savings (MoinAI)',fmtEuro(maC.savings),rel(maC.savings,maP.savings),'up',null,null,
    'Berechnung: Direkt vom Chatbot gelöste Anfragen × Kosten pro Chatbot-Konversation × 7. Kosten pro Konversation = 2.000 € MoinAI-Flat-Fee ÷ Chatbot-Konversationen der Woche. Quelle: Sheet MoinAI_KPIs, Spalte \'Weekly Savings\'.'),
  tile('Net Contribution (Refunds)',fmtEuro(rfC.net),rel(rfC.net,rfP.net),'up'),
  tile('Messages per Ticket',
    cur.msg_per_ticket!=null?cur.msg_per_ticket.toLocaleString('de-DE',{maximumFractionDigits:2}):'–',
    rel(cur.msg_per_ticket,prev.msg_per_ticket),'down')
);

const C={s1:css('--s1'),s2:css('--s2'),s3:css('--s3'),s4:css('--s4'),s5:css('--s5'),s6:css('--s6'),s7:css('--s7'),s8:css('--s8')};

/* HubSpot */
chartCard(document.getElementById('sec-hubspot'),{
  title:'Ticketvolumen',hint:'User- und Partner-Tickets, gestapelt',
  legend:[{name:'User',color:C.s1},{name:'Partner',color:C.s2}],
  table:{cols:['Zeitraum','User','Partner','Gesamt'],
    rows:hs.map(r=>[r.label,fmtN(r.user),fmtN(r.partner),fmtN(r.created)])},
  spec:{defaultView:'balken',labels:kwl,yFmt:fmtN,totalLabelLast:true,
    series:[{name:'User',color:C.s1,values:hs.map(r=>r.user)},
            {name:'Partner',color:C.s2,values:hs.map(r=>r.partner)}]}
});
chartCard(document.getElementById('sec-hubspot'),{
  title:'CSAT',hint:'Anteil zufriedener Bewertungen · graue Linie: 4-Perioden-Trend',
  table:{cols:['Zeitraum','CSAT'],rows:hs.map(r=>[r.label,fmtP(r.csat)])},
  spec:{defaultView:'linie',labels:kwl,yFmt:v=>Math.round(v*100)+' %',target:0.71,trend:true,
    series:[{name:'CSAT',color:C.s2,values:hs.map(r=>r.csat),fmt:fmtP}]}
});
chartCard(document.getElementById('sec-hubspot'),{
  title:'Median First Reply Time',hint:'Stunden bis zur ersten Antwort (Median) · graue Linie: 4-Perioden-Trend',
  table:{cols:['Zeitraum','MTFR (h)'],rows:hs.map(r=>[r.label,fmtH(r.mtfr)])},
  spec:{defaultView:'linie',labels:kwl,yFmt:v=>v+' h',target:24,trend:true,
    series:[{name:'MTFR',color:C.s1,values:hs.map(r=>r.mtfr),fmt:fmtH}]}
});
chartCard(document.getElementById('sec-hubspot'),{
  title:'Messages',hint:'Vom Team gesendete Nachrichten',
  table:{cols:['Zeitraum','Messages'],rows:hs.map(r=>[r.label,fmtN(r.messages)])},
  spec:{defaultView:'linie',labels:kwl,yFmt:fmtN,labelLast:true,
    series:[{name:'Messages',color:C.s5,values:hs.map(r=>r.messages),fmt:fmtN}]}
});

/* MoinAI */
chartCard(document.getElementById('sec-moinai'),{
  title:'Conversations',hint:'Aufgeteilt nach Kanal',
  legend:[{name:'HubSpot',color:C.s1},{name:'Chatbot',color:C.s2}],
  table:{cols:['Zeitraum','HubSpot','Chatbot','Gesamt'],
    rows:ma.map(r=>[r.label,fmtN(r.conv_hubspot),fmtN(r.conv_chatbot),fmtN(r.conv_total)])},
  spec:{defaultView:'balken',labels:kwl,yFmt:fmtN,totalLabelLast:true,
    series:[{name:'HubSpot',color:C.s1,values:ma.map(r=>r.conv_hubspot)},
            {name:'Chatbot',color:C.s2,values:ma.map(r=>r.conv_chatbot)}]}
});
chartCard(document.getElementById('sec-moinai'),{
  title:'Automation Rate',hint:'Anteil automatisch gelöster Anfragen',
  legend:[{name:'Chatbot-Konversationen',color:C.s1,type:'line'},
          {name:'Alle Tickets',color:C.s2,type:'line'}],
  table:{cols:['Zeitraum','Chatbot','Alle Tickets'],
    rows:ma.map(r=>[r.label,fmtP(r.auto_bot),fmtP(r.auto_all)])},
  spec:{defaultView:'linie',labels:kwl,yFmt:v=>Math.round(v*100)+' %',labelSeries:1,target:0.30,
    series:[{name:'Chatbot',color:C.s1,values:ma.map(r=>r.auto_bot),fmt:fmtP},
            {name:'Alle Tickets',color:C.s2,values:ma.map(r=>r.auto_all),fmt:fmtP}]}
});
chartCard(document.getElementById('sec-moinai'),{
  title:'Savings',hint:'Ersparnis durch Chatbot-Automatisierung',
  explain:'Berechnung: Direkt vom Chatbot gelöste Anfragen × Kosten pro Chatbot-Konversation × 7. Kosten pro Konversation = 2.000 € MoinAI-Flat-Fee ÷ Chatbot-Konversationen der Woche. Quelle: Sheet MoinAI_KPIs, Spalte \'Weekly Savings\'.',
  table:{cols:['Zeitraum','Savings'],rows:ma.map(r=>[r.label,fmtEuro(r.savings)])},
  spec:{defaultView:'balken',labels:kwl,yFmt:v=>fmtN(v),totalLabelLast:true,
    series:[{name:'Savings',color:C.s2,values:ma.map(r=>r.savings),fmt:fmtEuro}]}
});
chartCard(document.getElementById('sec-moinai'),{
  title:'Takeovers',hint:'Konversationen mit Übernahme durch das Team',
  table:{cols:['Zeitraum','Takeovers'],rows:ma.map(r=>[r.label,fmtN(r.takeovers)])},
  spec:{defaultView:'linie',labels:kwl,yFmt:fmtN,labelLast:true,
    series:[{name:'Takeovers',color:C.s2,values:ma.map(r=>r.takeovers),fmt:fmtN}]}
});

/* MoinAI ROI: Savings vs. Flat Fee (monatlich, nur intern) */
if(!META.public){
  const y=new Date().getFullYear();
  const savByMonth={};
  D.moinai.forEach(r=>{const m=MONTHS_DE[monthOfKW(r.kw)];
    if(r.savings!=null)savByMonth[m]=(savByMonth[m]||0)+r.savings;});
  const roiRows=co.filter(c=>c.moinai_fee&&savByMonth[c.month])
    .map(c=>({month:c.month,fee:c.moinai_fee,sav:savByMonth[c.month],
              roi:savByMonth[c.month]/c.moinai_fee}));
  Object.keys(savByMonth).forEach(m=>{
    if(!roiRows.some(r=>r.month===m)&&co.length){
      const fee=co[co.length-1].moinai_fee;
      if(fee)roiRows.push({month:m,fee:fee,sav:savByMonth[m],roi:savByMonth[m]/fee});
    }});
  if(roiRows.length){
    const lastR=roiRows[roiRows.length-1];
    chartCard(document.getElementById('sec-moinai'),{
      title:'Chatbot-ROI: Savings vs. Kosten',
      hint:'Monatliche Ersparnis vs. MoinAI Flat Fee · '+lastR.month+': '+
        lastR.roi.toLocaleString('de-DE',{maximumFractionDigits:1})+'× ROI',
      legend:[{name:'Savings',color:C.s2,type:'line'},{name:'Flat Fee',color:C.s6,type:'line'}],
      table:{cols:['Monat','Savings','Flat Fee','ROI'],
        rows:roiRows.map(r=>[r.month,fmtEuro(r.sav),fmtEuro(r.fee),
          r.roi.toLocaleString('de-DE',{maximumFractionDigits:1})+'×'])},
      spec:{defaultView:'linie',labels:roiRows.map(r=>r.month),yFmt:fmtN,
        series:[{name:'Savings',color:C.s2,values:roiRows.map(r=>r.sav),fmt:fmtEuro},
                {name:'Flat Fee',color:C.s6,values:roiRows.map(r=>r.fee),fmt:fmtEuro}]}
    });
  }
}

/* Team */
const agents=['Eli','Jeanine','Vivien','Jan'];
const agentColors=[C.s1,C.s2,C.s3,C.s5];
const notesTxt=(function(){
  const parts=[];
  tm.forEach(r=>agents.forEach(a=>{
    if(r[a]&&r[a].note)parts.push(a+' '+r.label+': '+r[a].note);}));
  return parts.length?'Hinweise: '+parts.join(' · '):null;
})();
chartCard(document.getElementById('sec-team'),{
  title:'Messages pro Agent',
  hint:'Gesendete Nachrichten',
  hoverInfo:notesTxt||undefined,
  legend:agents.map((a,i)=>({name:a,color:agentColors[i],type:'line'})),
  table:{cols:['Zeitraum',...agents],
    rows:tm.map(r=>[r.label,...agents.map(a=>fmtN(r[a].messages))])},
  spec:{defaultView:'linie',labels:tm.map(r=>r.label),yFmt:fmtN,
    series:agents.map((a,i)=>({name:a,color:agentColors[i],
      values:tm.map(r=>r[a].messages),fmt:fmtN}))}
});
(function(){
  const withHours=agents.filter((a,i)=>tm.some(r=>r[a]&&r[a].active_hours&&r[a].messages));
  if(!withHours.length)return;
  chartCard(document.getElementById('sec-team'),{
    title:'Messages pro Aktivstunde',
    hint:'Produktivität fair verglichen: Nachrichten ÷ aktive Stunden (nur Agents mit erfassten Aktivstunden)',
    legend:withHours.map(a=>({name:a,color:agentColors[agents.indexOf(a)],type:'line'})),
    table:{cols:['Zeitraum',...withHours],
      rows:tm.map(r=>[r.label,...withHours.map(a=>{
        const m=r[a].messages,h=r[a].active_hours;
        return (m&&h)?(m/h).toLocaleString('de-DE',{maximumFractionDigits:1}):'–';})])},
    spec:{defaultView:'linie',labels:tm.map(r=>r.label),
      yFmt:v=>v.toLocaleString('de-DE',{maximumFractionDigits:0}),
      series:withHours.map(a=>({name:a,color:agentColors[agents.indexOf(a)],
        values:tm.map(r=>(r[a].messages&&r[a].active_hours)?r[a].messages/r[a].active_hours:null),
        fmt:v=>v.toLocaleString('de-DE',{maximumFractionDigits:1})+' / h'}))}
  });
})();
chartCard(document.getElementById('sec-team'),{
  title:'Ø Handling Time pro Agent',hint:'Minuten pro Vorgang',
  legend:agents.map((a,i)=>({name:a,color:agentColors[i],type:'line'})),
  table:{cols:['Zeitraum',...agents],
    rows:tm.map(r=>[r.label,...agents.map(a=>fmtMin(r[a].aht))])},
  spec:{defaultView:'linie',labels:tm.map(r=>r.label),yFmt:v=>v+' min',
    series:agents.map((a,i)=>({name:a,color:agentColors[i],
      values:tm.map(r=>r[a].aht),fmt:fmtMin}))}
});

/* Refunds */
chartCard(document.getElementById('sec-refunds'),{
  title:'Refund-Tickets',hint:'Positive vs. abgelehnte Refund-Anfragen',
  legend:[{name:'Abgelehnt',color:C.s8},{name:'Erstattet',color:C.s2}],
  table:{cols:['Zeitraum','Abgelehnt','Erstattet','Gesamt','Decline Rate'],
    rows:rf.map(r=>[r.label,fmtN(r.negative),fmtN(r.positive),
      fmtN(r.refund_tickets),fmtP(r.decline_rate)])},
  spec:{defaultView:'balken',labels:kwl,yFmt:fmtN,totalLabelLast:true,
    series:[{name:'Abgelehnt',color:C.s8,values:rf.map(r=>r.negative)},
            {name:'Erstattet',color:C.s2,values:rf.map(r=>r.positive)}]}
});
chartCard(document.getElementById('sec-refunds'),{
  title:'Refund Decline Rate',hint:'Anteil abgelehnter Refund-Anfragen',
  table:{cols:['Zeitraum','Decline Rate'],rows:rf.map(r=>[r.label,fmtP(r.decline_rate)])},
  spec:{defaultView:'linie',labels:kwl,yFmt:v=>Math.round(v*100)+' %',target:0.90,trend:true,
    series:[{name:'Decline Rate',color:C.s8,values:rf.map(r=>r.decline_rate),fmt:fmtP}]}
});
chartCard(document.getElementById('sec-refunds'),{
  title:'Money Contribution',hint:'Durch abgelehnte Refunds gesichertes Geld',
  legend:[{name:'Netto',color:C.s8,type:'line'},{name:'Brutto',color:C.s3,type:'line'}],
  table:{cols:['Zeitraum','Netto','Brutto'],
    rows:rf.map(r=>[r.label,fmtEuro(r.net),fmtEuro(r.gross)])},
  spec:{defaultView:'linie',labels:kwl,yFmt:fmtN,
    series:[{name:'Netto',color:C.s8,values:rf.map(r=>r.net),fmt:fmtEuro},
            {name:'Brutto',color:C.s3,values:rf.map(r=>r.gross),fmt:fmtEuro}]}
});
chartCard(document.getElementById('sec-refunds'),{
  title:'Contribution pro Refund-Ticket',
  hint:'Netto gesichertes Geld je Refund-Anfrage (Effizienz der Ablehnungen)',
  table:{cols:['Zeitraum','Netto je Ticket'],
    rows:rf.map(r=>[r.label,(r.net&&r.refund_tickets)?fmtEuro2(r.net/r.refund_tickets):'–'])},
  spec:{defaultView:'linie',labels:kwl,yFmt:v=>v+' €',
    series:[{name:'Netto je Ticket',color:C.s8,
      values:rf.map(r=>(r.net&&r.refund_tickets)?r.net/r.refund_tickets:null),fmt:fmtEuro2}]}
});
chartCard(document.getElementById('sec-refunds'),{
  title:'Share of Refund Tickets',hint:'Anteil der Refund-Tickets am Gesamtvolumen',
  table:{cols:['Zeitraum','Anteil'],rows:rf.map(r=>[r.label,fmtP(r.share)])},
  spec:{defaultView:'linie',labels:kwl,yFmt:v=>Math.round(v*100)+' %',labelLast:true,
    series:[{name:'Anteil',color:C.s8,values:rf.map(r=>r.share),fmt:fmtP}]}
});

/* Costs */
const coL=co.map(r=>r.month);
chartCard(document.getElementById('sec-costs'),{
  title:'Cost per Ticket (monatlich)',
  hint:META.public?'Kosten pro Ticket inkl. Chatbot':'CST-Gesamtkosten ÷ Created Tickets',
  table:META.public
    ?{cols:['Monat','Tickets','CPT'],
      rows:co.map(r=>[r.month,fmtN(r.tickets),fmtEuro2(r.cpt)])}
    :{cols:['Monat','Kosten gesamt','Tickets','CPT'],
      rows:co.map(r=>[r.month,fmtEuro(r.total_cost),fmtN(r.tickets),fmtEuro2(r.cpt)])},
  spec:{defaultView:'linie',labels:coL,yFmt:v=>v+' €',target:2.5,
    series:[{name:'CPT',color:C.s3,values:co.map(r=>r.cpt),fmt:fmtEuro2}]}
});
if(!META.public){
chartCard(document.getElementById('sec-costs'),{
  title:'CST-Gesamtkosten (monatlich)',hint:'Team + MoinAI Flat Fee, brutto',
  table:{cols:['Monat','Gesamtkosten'],rows:co.map(r=>[r.month,fmtEuro(r.total_cost)])},
  spec:{defaultView:'balken',labels:coL,yFmt:fmtN,totalLabelLast:true,
    series:[{name:'Gesamtkosten',color:C.s3,values:co.map(r=>r.total_cost),fmt:fmtEuro}]}
});
}
} /* Ende renderAll */

/* ---------- Start & Live-Refresh ---------- */
const wrap=document.querySelector('.wrap');
// Zeitraum aus URL wiederherstellen + Umschalten verdrahten
const periodSel=document.getElementById('periodSel');
const urlPeriod=new URLSearchParams(location.search).get('zeitraum');
if(['woche','monat','jahr'].includes(urlPeriod))periodSel.value=urlPeriod;
periodSel.addEventListener('change',()=>{
  const u=new URL(location);u.searchParams.set('zeitraum',periodSel.value);
  history.replaceState(null,'',u);
  if(lastData)renderAll(lastData);
});
document.getElementById('kmTitleRow').addEventListener('click',()=>{
  window.__kmOpen=!window.__kmOpen;
  document.getElementById('kmtiles').style.display=window.__kmOpen?'':'none';
  document.getElementById('kmTitleRow').classList.toggle('open',!!window.__kmOpen);
});
renderAll(DATA); // eingebettete Daten sofort zeigen — nie eine leere Seite

async function refreshLive(){
  const btn=document.getElementById('reloadBtn');
  btn.classList.add('spin');
  wrap.style.opacity='.55'; // vorherige Ansicht halten, kein Layout-Sprung
  try{
    const D=await fetchLive();
    renderAll(D);
  }catch(e){
    console.warn('Live-Daten nicht erreichbar, zeige letzten Build:',e);
  }finally{
    wrap.style.opacity='';
    btn.classList.remove('spin');
  }
}
document.getElementById('reloadBtn').addEventListener('click',refreshLive);
refreshLive(); // beim Öffnen automatisch aktuelle Zahlen laden
</script>
</body>
</html>
"""


def main():
    xlsx, out = sys.argv[1], sys.argv[2]
    public = '--public' in sys.argv[3:]
    data = extract(xlsx)
    if public:
        # Vertrauliche Kostendaten aus der geteilten Version entfernen
        for c in data['costs']:
            c['total_cost'] = None
    import os
    meta = {'source': os.path.basename(xlsx),
            'updated': date.today().strftime('%d.%m.%Y'),
            'public': public}
    html = HTML_TEMPLATE.replace('__DATA__', json.dumps(data)) \
                        .replace('__META__', json.dumps(meta))
    with open(out, 'w') as f:
        f.write(html)
    print(f'OK: {out} geschrieben ({len(html)} Bytes), '
          f'KW {data["hubspot"][0]["kw"]}–{data["hubspot"][-1]["kw"]}')


if __name__ == '__main__':
    main()
